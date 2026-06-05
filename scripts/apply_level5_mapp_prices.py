"""
apply_level5_mapp_prices.py

Reads MAPP prices from 'Order Input - Level5' sheet and writes them as
Regular price for every Level5 product in the catalog CSV.

Rules:
  - simple / variation rows  → set Regular price directly from MAPP table
  - variable (parent) rows   → set Regular price to the MIN of matched child prices
  - Any SKU with no match is reported but left untouched
"""

import csv
import openpyxl
from collections import defaultdict

XLSX  = r'products\scraped_results\brands\Level5\level5_parts_order_form.xlsx'
CSV   = r'products\Production\launch\dtb_woocommerce_official_catalog.csv'
SHEET = 'Order Input - Level5'

# ── 1. Load MAPP prices from Excel ───────────────────────────────────────────
def load_mapp_prices(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    prices = {}
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        sku  = row[1]   # column B
        mapp = row[7]   # column H
        if sku is None or mapp is None:
            continue
        sku_str = str(sku).strip()
        try:
            prices[sku_str] = round(float(mapp), 2)
        except (ValueError, TypeError):
            pass
    return prices

# ── 2. Load catalog ──────────────────────────────────────────────────────────
def load_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return reader.fieldnames, list(reader)

# ── 3. Apply prices ──────────────────────────────────────────────────────────
def apply_prices(fieldnames, rows, mapp):
    # Build parent → [child prices] for variable parents
    parent_prices = defaultdict(list)
    for row in rows:
        if row.get('Meta: _dtb_brand_key', '').strip().lower() != 'level5':
            continue
        sku = row.get('SKU', '').strip()
        if row.get('Type', '') == 'variation' and sku in mapp:
            parent_sku = row.get('Parent', '').strip()
            parent_prices[parent_sku].append(mapp[sku])

    changed   = []
    unmatched = []

    for row in rows:
        if row.get('Meta: _dtb_brand_key', '').strip().lower() != 'level5':
            continue

        sku  = row.get('SKU', '').strip()
        ptype = row.get('Type', '').strip()

        if ptype == 'variable':
            # Set to min child price if we have any matched children
            child_prices = parent_prices.get(sku, [])
            if child_prices:
                price = round(min(child_prices), 2)
                row['Regular price'] = str(price)
                changed.append((sku, ptype, price, 'min of children'))
            else:
                # Try direct match (some variable parents share a SKU with their Excel entry)
                if sku in mapp:
                    row['Regular price'] = str(mapp[sku])
                    changed.append((sku, ptype, mapp[sku], 'direct'))
                else:
                    unmatched.append((sku, ptype))

        elif ptype in ('simple', 'variation'):
            if sku in mapp:
                row['Regular price'] = str(mapp[sku])
                changed.append((sku, ptype, mapp[sku], 'direct'))
            else:
                unmatched.append((sku, ptype))

    return changed, unmatched

# ── 4. Write catalog ─────────────────────────────────────────────────────────
def write_csv(path, fieldnames, rows):
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    mapp = load_mapp_prices(XLSX, SHEET)
    print(f'Loaded {len(mapp)} MAPP prices from Excel\n')

    fieldnames, rows = load_csv(CSV)
    changed, unmatched = apply_prices(fieldnames, rows, mapp)
    write_csv(CSV, fieldnames, rows)

    print(f'{"SKU":<40} {"Type":<12} {"Price":>10}  Source')
    print('-' * 78)
    for sku, ptype, price, src in sorted(changed, key=lambda x: (x[1], x[0])):
        print(f'{sku:<40} {ptype:<12} {price:>10}  {src}')

    print(f'\n✓  {len(changed)} prices applied')

    if unmatched:
        print(f'\n⚠  {len(unmatched)} Level5 SKUs with no MAPP match (price unchanged):')
        for sku, ptype in sorted(unmatched):
            print(f'   {sku}  ({ptype})')
