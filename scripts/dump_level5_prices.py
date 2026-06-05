import openpyxl

wb = openpyxl.load_workbook(
    r'products\scraped_results\brands\Level5\level5_parts_order_form.xlsx',
    data_only=True
)
ws = wb['Order Input - Level5']

print("{:<15} {:>10}  {}".format("SKU", "MAPP", "Description"))
print("-" * 70)
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
    sku  = row[1]
    mapp = row[7]
    desc = row[2]
    if sku is not None and mapp is not None:
        print("{:<15} {:>10}  {}".format(str(sku).strip(), str(mapp), str(desc) if desc else ""))
